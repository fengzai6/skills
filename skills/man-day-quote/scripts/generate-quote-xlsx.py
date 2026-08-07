#!/usr/bin/env python3
"""Generate man-day quote xlsx from JSON rows.

Template resolution:
  1) --template PATH if provided and exists
  2) code-built blank sheet (headers + SUM formula) — no external xlsx required

JSON input (file or stdin):
{
  "title": "可选，仅用于日志",
  "rows": [
    {"seq": 1, "feature": "FR-01", "type": "邏輯", "detail": "……", "days": 0.5},
    ...
  ]
}

Columns: A序号 B功能 C类型 D报价明细 E人天 F(空) G Total(SUM)
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ModuleNotFoundError as exc:
    if exc.name != "openpyxl":
        raise
    raise SystemExit(
        "missing dependency: openpyxl\n"
        "install with: python3 -m pip install openpyxl"
    ) from exc

HEADERS = ["序号", "功能", "类型", "报价明细", "人天", None, "Total"]
COL_SEQ, COL_FEATURE, COL_TYPE, COL_DETAIL, COL_DAYS = 1, 2, 3, 4, 5
COL_TOTAL = 7  # G

# Same list as 模板人天报价.xlsx type dropdown (C column)
TYPE_OPTIONS = (
    "邏輯",
    "軟件接口對接",
    "頁面",
    "定義一個實體（含CURD）",
    "定義一個實體的DBUP",
    "接口",
    "硬件接口對接",
    "硬件集成",
    "Discount",
)
TYPE_ALIASES = {
    "逻辑": "邏輯",
    "软件接口对接": "軟件接口對接",
    "页面": "頁面",
    "定义一个实体（含CURD）": "定義一個實體（含CURD）",
    "定义一个实体的DBUP": "定義一個實體的DBUP",
    "硬件接口对接": "硬件接口對接",
    "硬件集成": "硬件集成",
}

# Align with 模板人天报价.xlsx / DAOTODO quotes (F spacer ≈ E/G, not hairline)
COLUMN_WIDTHS = {
    1: 8,  # 序号
    2: 18,  # 功能
    3: 26,  # 类型
    4: 56,  # 报价明细
    5: 12,  # 人天
    6: 12,  # 空列（与人天/Total 同量级，避免挤成一条缝）
    7: 12,  # Total
}

CENTER = Alignment(horizontal="center", vertical="center")
# 报价明细：垂直居中、水平左对齐，便于长句阅读
DETAIL_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def apply_column_widths(ws) -> None:
    for col, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def apply_sheet_style(ws) -> None:
    """表头全居中；数据行除报价明细外全居中。"""
    max_row = max(ws.max_row, 2)
    for row in range(1, max_row + 1):
        for col in range(1, 8):
            cell = ws.cell(row, col)
            if row == 1 or col != COL_DETAIL:
                cell.alignment = CENTER
            else:
                cell.alignment = DETAIL_ALIGN


def apply_type_dropdown(ws) -> None:
    """C2:C 下拉，与官方模板一致；先清旧的 C 列 list 校验再加，避免重复。"""
    keep = []
    for dv in ws.data_validations.dataValidation:
        refs = str(dv.sqref)
        # drop validations that only/primarily target type col C
        if refs.startswith("C") or " C" in f" {refs}":
            continue
        keep.append(dv)
    ws.data_validations.dataValidation = keep

    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(TYPE_OPTIONS) + '"',
        allow_blank=True,
        showDropDown=False,  # False = show dropdown arrow (openpyxl quirk)
        showErrorMessage=True,
    )
    dv.add("C2:C1048576")
    ws.add_data_validation(dv)


def build_blank_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i, h in enumerate(HEADERS, start=1):
        if h is not None:
            ws.cell(1, i, h)
    ws.cell(2, COL_TOTAL, "=SUM(E:E)")
    apply_column_widths(ws)
    apply_sheet_style(ws)
    apply_type_dropdown(ws)
    return wb


def load_template(path: Path | None) -> Workbook:
    if path is not None:
        if not path.is_file():
            raise SystemExit(f"template not found: {path}")
        return load_workbook(path)
    return build_blank_workbook()


def clear_data_rows(ws) -> None:
    """Keep header row; clear old data so re-fill is clean."""
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def normalize_type(value, row_num: int) -> str:
    typ = str(value or "").strip()
    typ = TYPE_ALIASES.get(typ, typ)
    if typ not in TYPE_OPTIONS:
        allowed = " | ".join(TYPE_OPTIONS)
        raise SystemExit(f"row {row_num}: invalid type {typ!r}; allowed: {allowed}")
    return typ


def normalize_days(value, row_num: int) -> float:
    if value in (None, ""):
        raise SystemExit(f"row {row_num}: days is required")
    try:
        days = float(value)
    except (TypeError, ValueError) as exc:
        raise SystemExit(f"row {row_num}: invalid days {value!r}") from exc
    if days <= 0:
        raise SystemExit(f"row {row_num}: days must be positive")
    if not (days * 2).is_integer():
        raise SystemExit(f"row {row_num}: days must use 0.5 increments")
    return days


def normalize_rows(rows: list[dict]) -> list[dict]:
    normalized = []
    for seq, item in enumerate(rows, start=1):
        if not isinstance(item, dict):
            raise SystemExit(f"row {seq}: expected object")
        detail = str(item.get("detail") or "").strip()
        if not detail:
            raise SystemExit(f"row {seq}: detail is required")
        normalized.append(
            {
                "seq": seq,
                "feature": str(item.get("feature") or item.get("fr") or "").strip(),
                "type": normalize_type(item.get("type"), seq),
                "detail": detail,
                "days": normalize_days(item.get("days"), seq),
            }
        )
    return normalized


def write_rows(ws, rows: list[dict]) -> None:
    clear_data_rows(ws)
    # ensure header
    for i, h in enumerate(HEADERS, start=1):
        if h is not None and not ws.cell(1, i).value:
            ws.cell(1, i, h)
        if h is None:
            ws.cell(1, i, None)

    for i, item in enumerate(rows, start=1):
        r = i + 1  # data starts row 2
        ws.cell(r, COL_SEQ, item["seq"])
        ws.cell(r, COL_FEATURE, item["feature"])
        ws.cell(r, COL_TYPE, item["type"])
        ws.cell(r, COL_DETAIL, item["detail"])
        ws.cell(r, COL_DAYS, item["days"])

    # Total formula on G2 (match historical quotes)
    ws.cell(2, COL_TOTAL, "=SUM(E:E)")
    if ws.cell(1, COL_TOTAL).value in (None, ""):
        ws.cell(1, COL_TOTAL, "Total")
    apply_column_widths(ws)
    apply_sheet_style(ws)
    apply_type_dropdown(ws)


def main() -> None:
    ap = argparse.ArgumentParser(description="Generate man-day quote xlsx")
    ap.add_argument(
        "-i",
        "--input",
        help="JSON file path; default stdin",
    )
    ap.add_argument(
        "-o",
        "--output",
        required=True,
        help="Output xlsx path",
    )
    ap.add_argument(
        "-t",
        "--template",
        help="Optional user template xlsx (preferred when provided)",
    )
    ap.add_argument(
        "--blank",
        action="store_true",
        help="Write blank template only (ignore rows)",
    )
    args = ap.parse_args()

    template_path = Path(args.template).expanduser() if args.template else None
    out = Path(args.output).expanduser()
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = load_template(template_path)
    ws = wb.active

    if args.blank:
        clear_data_rows(ws)
        for i, h in enumerate(HEADERS, start=1):
            if h is not None:
                ws.cell(1, i, h)
        ws.cell(2, COL_TOTAL, "=SUM(E:E)")
        apply_column_widths(ws)
        apply_sheet_style(ws)
        apply_type_dropdown(ws)
        wb.save(out)
        print(str(out))
        return

    raw = Path(args.input).read_text(encoding="utf-8") if args.input else sys.stdin.read()
    data = json.loads(raw)
    if isinstance(data, list):
        rows = data
    else:
        rows = data.get("rows") or []
    if not rows:
        raise SystemExit("no rows in input JSON")

    rows = normalize_rows(rows)
    write_rows(ws, rows)
    wb.save(out)
    total = sum(r["days"] for r in rows)
    print(str(out))
    print(f"rows={len(rows)} total_days={total:g}", file=sys.stderr)


if __name__ == "__main__":
    main()
